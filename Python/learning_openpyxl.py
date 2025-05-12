import openpyxl

def main():
    inv_file = openpyxl.load_workbook("/Users/snakkina/Workspace/Python/Datafiles/inventory.xlsx")
    product_list = inv_file["Sheet1"]
    calculations(product_list)
    inv_file.save("/Users/snakkina/Workspace/Python/Datafiles/inventory_with_total_value.xlsx")


def calculations(product_list):
    products_per_supplier = {}
    inventory_value_per_supplier = {}
    for product_row in range(2, product_list.max_row + 1):
        supplier_name = product_list.cell(product_row, 4).value
        price = product_list.cell(product_row, 3).value
        inventory = product_list.cell(product_row, 2).value
        inventory_price = product_list.cell(product_row, 5)

        #select count(PRODUCT_NO) group by SUPPLIER
        if supplier_name in products_per_supplier:
            products_per_supplier[supplier_name] = products_per_supplier[supplier_name] + 1
            #print("Supplier Updated")
        else:
            products_per_supplier[supplier_name] = 1
            #print("Supplier Added")

        #select sum(PRICE) group by SUPPLIER
        if supplier_name in inventory_value_per_supplier:
            inventory_value_per_supplier[supplier_name] = inventory_value_per_supplier[supplier_name] + (inventory * price)
        else:
            inventory_value_per_supplier[supplier_name] = inventory * price
        
        #select INVENTORY * PRICE for every SUPPLIER
        inventory_price.value = inventory * price

    print(products_per_supplier)
    print(inventory_value_per_supplier)

main()